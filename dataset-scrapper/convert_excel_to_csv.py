# pip3 install pandas
# pip3 install openpyxl

import openpyxl
import pandas as pd


EXCEL_FILENAME = "updating_music_info.xlsx"
EXCEL_FILENAME2 = "updating_music_info2.xlsx"

def convert_to_csv(excel_filename):
    wb_obj = openpyxl.load_workbook(excel_filename)
    excel_data = []
    for worksheet_name in wb_obj.sheetnames:
        print("*"*30)
        print(worksheet_name)
        sheet_obj = wb_obj[worksheet_name]
        
        max_row = sheet_obj.max_row
        max_col = sheet_obj.max_column

        for line in range(2, max_row + 1):
            line_data = []
            for col in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row = line, column = col)
                line_data.append(cell_obj.value)
                # print(cell_obj.value)
            print(line_data)
            excel_data.append(line_data)
    return excel_data
    
data = convert_to_csv(EXCEL_FILENAME)    
data.extend(convert_to_csv(EXCEL_FILENAME2))

lyrics = pd.DataFrame(data, columns=['artist', 'year', 'song', 'album', 'lyrics'])
lyrics.to_csv('lyrics.csv')


df = pd.read_csv ('lyrics.csv', index_col=0)
print(df)