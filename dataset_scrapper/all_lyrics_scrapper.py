# pip3 install requests
# pip3 install beautifulsoup4
# pip3 install openpyxl

import requests
from bs4 import BeautifulSoup
from bs4 import Comment
import json
import openpyxl

EXCEL_FILENAME = "music_info.xlsx"
URL = "https://www.azlyrics.com/"
agent = {"User-Agent":'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36'}

tabs_class = "btn btn-menu"
artists_class = "col-sm-6 text-center artist-col"
song_class = "listalbum-item"
album_class = "songinalbum_title"
# # => 19


disclaimer = "Sorry about that"


def get_page(url):
    page = requests.get(url, headers=agent)
    return BeautifulSoup(page.content, "lxml")

def get_artist_songs(artist_element):
    artist_data = []
    artist_soup = get_page(URL + artist_element['href'])
    # print(artist_element.text) 
    songs_divs = artist_soup.find_all("div", class_=song_class)
    for div in songs_divs:
        song_elem = div.find("a" , href=lambda href: href and href.startswith("../"))
        print(song_elem['href'])
        song_soup = get_page(song_elem['href'].replace("../", URL))

        album_year_info = song_soup.find("div", class_=album_class).text
        # print(album_year_info) # eg: album: "A Fever You Can't Sweat Out" (2005)
        song_info = {
            "artist": artist_element.text,
            "year": album_year_info.replace("album:", "").split("(")[1].split(")")[0],
            "song": song_elem.text,
            "album": album_year_info.replace("album:", "").split("(")[0].strip(),
            "lyrics": get_lyrics(song_soup)
        }
        print(song_info)
        artist_data.append(song_info)
    return artist_data
        


def get_lyrics(song_soup):
    for comment in song_soup.find_all(text=lambda text: isinstance(text, Comment)):
        if disclaimer in comment:
            return comment.parent.text
        

def get_tabs(URL):
    soup = get_page(URL)
    return soup.find_all("a", class_=tabs_class, href=True)




def parse_tabs(tabs_elements):
    data = []
    for tab_element in tabs_elements:
        
        print(tab_element['href'], end="\n"*2)

        artists_pages = extract_artists_from_tab(tab_element)
        
        for artist in artists_pages:
            data += get_artist_songs(artist)
    write_excel(data)
        

def extract_artists_from_tab(tab_element):
    tab_text = tab_element.text.lower()
    if tab_text == "#":
        tab_text = "19"
    tab_soup = get_page("https:" + tab_element['href'])
    # print(tab_soup)
    artists_column = tab_soup.find_all("div", {"class": artists_class})
    artists = []
    for artist_column in artists_column:
        # some artist appear on multiple pages
        artist_elements = artist_column.find_all("a" , href=lambda href: href and href.startswith(tab_text))
        for artist_element in artist_elements:
            artists.append(artist_element)
            # print(URL + artist_element['href'])
    return artists

def write_excel(data):
    wb_obj = openpyxl.load_workbook(EXCEL_FILENAME)
    sheet = wb_obj.active

    headers = [ "artist", "year", "song", "album", "lyrics"]
    for idx, header in enumerate(headers):
        sheet.cell(row=1, column=idx+1).value = header

    for line, song in enumerate(data):
        for idx, val in enumerate(song.values()):
            sheet.cell(row=line+2, column=idx+1).value = val

    
    wb_obj.save(EXCEL_FILENAME)




################################################################ 
tabs_elements = get_tabs(URL)
print(tabs_elements)
parse_tabs(tabs_elements)



