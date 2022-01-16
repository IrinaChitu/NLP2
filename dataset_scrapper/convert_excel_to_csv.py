# pip3 install pandas
# pip3 install openpyxl
# pip3 install spacy-langdetect
# pip3 install spacy
# python3 -m spacy download en_core_web_sm

import openpyxl
import pandas as pd
import spacy
from spacy.language import Language
from spacy_langdetect import LanguageDetector
from collections import Counter

EXCEL_FILENAME = "updating_music_info_en_only.xlsx"
CSV_FILENAME = "dataset_manual.csv"
NEW_LINE_CELL_CHAR = "_x000D_"

def get_lang_detector(nlp, name):
    return LanguageDetector()
    

def clean_up(lyrics):
    return lyrics.replace("\n", " ").replace("_x000D_", "").replace("(", " ").replace(")", " ").replace("?", " ")

def predict_lang(lyrics):
    nlp = spacy.load("en_core_web_sm")
    Language.factory("language_detector", func=get_lang_detector)
    nlp.add_pipe('language_detector', last=True)
    doc = nlp(clean_up(lyrics))

    pred_lang = doc._.language
    if pred_lang["score"] < 0.60:
        scores = []
        for i in range(7):
            scores.append(doc._.language["language"])
        return Counter(scores).most_common(1)[0][0]
    else:
        return pred_lang["language"]



def convert_to_csv(excel_filename):
    wb_obj = openpyxl.load_workbook(excel_filename)
    excel_data = []

    for worksheet_name in wb_obj.sheetnames:
        # print("*"*30)
        # print(worksheet_name)
        sheet_obj = wb_obj[worksheet_name]
        
        max_row = sheet_obj.max_row
        max_col = sheet_obj.max_column

        for line in range(2, max_row + 1):
            line_data = []
            for col in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row = line, column = col)
                cell_data = cell_obj.value
                # if cell_data is not None:
                #     cell_data = cell_data.replace("\"", "")
                line_data.append(cell_data)
                # print(cell_obj.value)
            if line_data == [] or line_data[0] is None or line_data[0] == "":
                print(line_data)
                print("line_data", line)

            if line_data[4] is not None and line_data[4] != '':
                line_data[4] = line_data[4].replace(NEW_LINE_CELL_CHAR, "")
                lang = predict_lang(line_data[4])
                if lang == "en":
                    excel_data.append(line_data)
                else:
                    print(line_data[0], line_data[2], lang)
            else:
                print(line_data[0], line_data[2], "empty line")
                
    return excel_data
    
data = convert_to_csv(EXCEL_FILENAME)    

lyrics = pd.DataFrame(data, columns=['artist', 'year', 'song', 'album', 'lyrics'])
lyrics.to_csv(CSV_FILENAME)


df = pd.read_csv (CSV_FILENAME, index_col=0)
print(df)