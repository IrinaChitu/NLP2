# pip3 install requests
# pip3 install beautifulsoup4
# pip3 install openpyxl

ARTISTS = [
"/a/abba.html",
"/a/acdc.html",
"/a/adele.html",
"/a/akon.html",
"/a/alanwalker.html",
"/a/alecbenjamin.html",
"/a/alltimelow.html",
"/a/arianagrande.html",
"/a/avicii.html",
"/a/awolnation.html",
"/b/bsb.html",
"/b/barbrastreisand.html",
"/b/bastille.html",
"/b/beatles.html",
"/d/depeche.html",
"/d/dion.html",
"/d/dylan.html",
"/e/edsheeran.html",
"/e/elvis.html",
"/e/eminem.html",
"/g/guns.html",
"/j/jackson.html",
"/l/labrinth.html",
"/l/lavigne.html",
"/l/lorde.html",
"/m/madonna.html",
"m/maroon5.html",
"/m/metallica.html",
"/m/muse.html",
"/n/nickiminaj.html",
"/p/panicatthedisco.html",
"/p/paulmccartney.html",
"/p/postmalone.html",
"/q/queen.html",
"/s/samsmith.html",
"/s/selenagomez.html",
"/t/timberlake.html",
"/u/usher.html"
]

import requests
from bs4 import BeautifulSoup
from bs4 import Comment
import json
import openpyxl
import time 


EXCEL_FILENAME = "updating_music_info.xlsx"
LOGS_FILENAME = "logs.txt"
URL = "https://www.azlyrics.com/"
agent = {"User-Agent":'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36'}

tabs_class = "btn btn-menu"
artists_class = "col-sm-6 text-center artist-col"
song_class = "listalbum-item"
album_class = "songinalbum_title"
albums_id = "listAlbum"
# # => 19


disclaimer = "Sorry about that"


def get_page(url, delay_time, message):
    page = requests.get(url, headers=agent)

    print("Waiting {0}s after fetching {1}".format(delay_time, message) )
    time.sleep(delay_time)
    
    return BeautifulSoup(page.content, "lxml")

def extract_album_info(album_year_info):
    album_year, album_name = "", ""
    if album_year_info is not None:
        if "(" in album_year_info and ")" in album_year_info:
            album_year_info = album_year_info.replace("album:", "")

            idx_l = album_year_info.rfind("(")
            idx_r = album_year_info.rfind(")")
            album_year = album_year_info[idx_l+1:idx_r]
            album_name = album_year_info[:idx_l].strip()

    return album_year, album_name

def get_artist_songs(artist_url):
    artist_data = []
    artist_soup = get_page(artist_url, 10, "artist page")
    artist_name = artist_soup.find('title').string
    if "Lyrics" in artist_name:
        artist_name = artist_name.replace(" Lyrics", "")
    
    albums_div = artist_soup.find("div", {"id": albums_id})

    div_children = albums_div.findChildren("div" , recursive=False)
    album_year_info, album_year, album_name = "", "", ""
    
    for idx, child in enumerate(div_children):

        if child.has_attr("class") and child["class"][0] == "album": 
            if album_year_info != child.text:
                print("*"*30)
                album_year_info = child.text
                print(album_year_info)
            
            album_year, album_name = extract_album_info(album_year_info)

        elif child.has_attr("class") and child["class"][0] == "listalbum-item":
            song_name = child.text
    
            print("div listalbum-item", child)
            starts_with = "/lyrics/"
            song_elem = child.find("a")
            if song_elem is not None:
                if URL in song_elem['href']:
                    song_url = song_elem['href']
                else:
                    song_url = URL + song_elem['href']
                song_soup = get_page(song_url, 10, "song page")
                print("song url", song_url)

                song_info = {
                    "artist": artist_name,
                    "year": album_year,
                    "song": song_name,
                    "album": album_name,
                    "lyrics": get_lyrics(song_soup, URL + song_elem['href'])
                }
                print(song_info)
                artist_data.append(song_info)
    return artist_data, artist_name


def get_lyrics(song_soup, song_url):
    for comment in song_soup.find_all(text=lambda text: isinstance(text, Comment)):
        if disclaimer in comment:
            return comment.parent.text
    

def write_excel(data, worksheet_name):
    wb_obj = openpyxl.load_workbook(EXCEL_FILENAME)
    if worksheet_name in wb_obj.sheetnames:
        sheet = wb_obj[worksheet_name]
    else:
        wb_obj.create_sheet(worksheet_name)
        sheet = wb_obj[worksheet_name]
        headers = [ "artist", "year", "song", "album", "lyrics"]
        for idx, header in enumerate(headers):
            sheet.cell(row=1, column=idx+1).value = header
    
    for line, song in enumerate(data):
        for idx, val in enumerate(song.values()):
            sheet.cell(row=line+2, column=idx+1).value = val

    
    wb_obj.save(EXCEL_FILENAME)



for artist_path in ARTISTS:
    artist_data, artist_name = get_artist_songs(URL + artist_path)
    write_excel(artist_data, artist_name)


